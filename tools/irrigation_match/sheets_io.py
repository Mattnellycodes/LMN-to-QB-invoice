"""Read and write the Google Sheets, or read local .xlsx snapshots for offline runs.

Live access goes through the Apps Script web app in apps_script/Code.gs,
deployed under the sheet owner's Google account. The deployment URL and shared
token come from the APPS_SCRIPT_URL and APPS_SCRIPT_TOKEN environment
variables — no GCP project or service account involved.
"""

from __future__ import annotations

import os

Rows = list[list[object]]


class SnapshotBackend:
    """Reads tabs from local .xlsx files; writes are reported, not performed."""

    def __init__(self, schedule_path: str, irrigation_path: str):
        from openpyxl import load_workbook

        self._books = {
            "schedule": load_workbook(schedule_path, read_only=True),
            "irrigation": load_workbook(irrigation_path, read_only=True),
        }

    def read_tab(self, sheet: str, tab: str) -> Rows:
        book = self._books[sheet]
        if tab not in book.sheetnames:
            return []
        return [list(row) for row in book[tab].iter_rows(values_only=True)]

    def replace_tab(self, sheet: str, tab: str, rows: Rows) -> None:
        print(
            f"[snapshot] would replace '{tab}' with {len(rows)} rows (header included)"
        )

    def append_rows(self, sheet: str, tab: str, rows: Rows) -> None:
        print(f"[snapshot] would append {len(rows)} rows to '{tab}'")

    def flush(self) -> None:
        pass


class AppsScriptBackend:
    """Live access through the Apps Script web app (see apps_script/Code.gs).

    All readable tabs arrive in one GET; writes are batched into one POST by
    main() calling replace_tab/append_rows, which stage, and flush() sending.
    """

    def __init__(self) -> None:
        self._url = os.environ.get("APPS_SCRIPT_URL")
        self._token = os.environ.get("APPS_SCRIPT_TOKEN")
        if not self._url or not self._token:
            raise RuntimeError("APPS_SCRIPT_URL and APPS_SCRIPT_TOKEN must be set")
        self._cache: dict[str, dict[str, Rows]] | None = None
        self._pending: dict[str, dict[str, Rows]] = {"replace": {}, "append": {}}

    def read_tab(self, sheet: str, tab: str) -> Rows:
        if self._cache is None:
            self._cache = self._request_json("get")
        return self._cache.get(sheet, {}).get(tab, [])

    def replace_tab(self, sheet: str, tab: str, rows: Rows) -> None:
        self._assert_schedule(sheet)
        self._pending["replace"][tab] = rows

    def append_rows(self, sheet: str, tab: str, rows: Rows) -> None:
        if not rows:
            return
        self._assert_schedule(sheet)
        self._pending["append"][tab] = rows

    def flush(self) -> None:
        if not self._pending["replace"] and not self._pending["append"]:
            return
        response = self._request_json("post", self._pending)
        if not response.get("ok"):
            raise RuntimeError(f"apps script write failed: {response}")
        print(f"apps script wrote: {response.get('written')}")
        self._pending = {"replace": {}, "append": {}}

    def _request_json(self, method: str, body: dict | None = None) -> dict:
        import requests

        params = {"token": self._token}
        if method == "get":
            response = requests.get(self._url, params=params, timeout=120)
        else:
            response = requests.post(self._url, params=params, json=body, timeout=120)
        response.raise_for_status()
        data = response.json()
        if isinstance(data, dict) and data.get("error"):
            raise RuntimeError(f"apps script error: {data['error']}")
        return data

    @staticmethod
    def _assert_schedule(sheet: str) -> None:
        if sheet != "schedule":
            raise ValueError("writes are only allowed on the schedule master")
